#!/usr/bin/env python3

""""description

.. note:: This program is free software; you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation; either version 2 of the License, or
    (at your option) any later version.

"""

__author__ = 'Luke Coburn'
__date__ = '2020-11-23'
__copyright__ = 'Copyright 2020'

import os.path
import pickle
import time
from urllib.parse import quote_plus

import qgis
from qgis.core import *
from qgis.gui import *
from qgis.PyQt.QtCore import *
from qgis.PyQt.QtGui import *
from qgis.PyQt.QtWidgets import *

from shapely.geometry import Polygon
from shapely.ops import cascaded_union

from dialogs.loggerHelpDialog import loggerHelpDialog
from dialogs.ProjectLoadDialog import ProjectLoadDialog
from dialogs.ManualMeasuresInfoDialog import ManualMeasuresInfoDialog
from dialogs.ManualRulesDialog import ManualRulesDialog
from dialogs.ManualExtensionsInfoDialog import ManualExtensionsInfoDialog
from dialogs.ManualLoftInfoDialog import ManualLoftInfoDialog
from dialogs.ManualComboInfoDialog import ManualComboInfoDialog
from dialogs.ManualURLInfoDialog import ManualURLInfoDialog
from dialogs.ManualCheckURLInfoDialog import ManualCheckURLInfoDialog
from dialogs.HouseInfoDialog import HouseInfoDialog
from dialogs.SplitHouseDialog import SplitHouseDialog
from dialogs.ManualTileDialog import ManualTileDialog
from dialogs.ManualLoftTileDialog import ManualLoftTileDialog
from dialogs.ReportInfoDialog import ReportInfoDialog
from dialogs.ShowManualMeasures import ShowManualMeasures
from dialogs.ShowLanesImage import ShowLanesImage
from dialogs.ShowExtensions import ShowExtensions
from dialogs.ShowLoftImage import ShowLoftImage
from dialogs.ManualPlannerNotesDialog import ManualPlannerNotesDialog

from roofus.extensions import Extensions

from utils.mapTools import *
from ui_mainWindow import Ui_MainWindow
from utils.Utils import *
import xlsxwriter
import datetime
import openpyxl
import requests
import re

from pathlib import Path
import sys

from webscraper.utils_ws.propertydata_scraper import property_data_scraper

home = str(Path.home())

# sys.path.append(path.abspath(home))
sys.path.append('/MyComputer')
# sys.path.append('/MyComputer/geometryTools')
# sys.path.append('/MyComputer/lanu_reader')
# sys.path.append('/MyComputer/lanu_webscraper')

from lanuApps.geometryTools.geometry import Geometry
from lanuApps.roofus.reader import Reader, Box, Block

from lanuApps.webscraper.utils_ws.constants import *
from lanuApps.webscraper.scripts.process_house_input import ProcessSingleHouse
from lanuApps.webscraper.utils_ws.useful_functions import UsefulFunctions
from lanuApps.webscraper.utils_ws.propertydata_scraper import *

logging.basicConfig(level=logging.DEBUG if DEBUG else logging.WARNING)
logger = logging.getLogger(__file__)

os.environ['DYLD_FRAMEWORK_PATH'] = "/QGIS/resources"
os.environ['QGIS_PREFIX'] = "/usr"


class HouseBounds(QMainWindow, Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.now = time.time()
        self.editing = False
        self.modified = False
        self.scrapeNotDone = True
        self.setupUi(self)

        self.actionLogger.triggered.connect(self.logDialog)
        self.actionQuit.triggered.connect(self.close)
        self.actionOpenProject.triggered.connect(self.openProjectDialog)
        self.actionSatelliteBaseMap.triggered.connect(self.showVisibleMapLayers)
        self.showExt.triggered.connect(self.showExtensionImage)
        self.showMM.triggered.connect(self.showManualMeasures)
        self.showLanes.triggered.connect(self.showLanesImage)
        self.showLoft.triggered.connect(self.showLoftImage)
        self.actionShowBasemap1Layer.triggered.connect(self.showBasemap1Layer)
        self.actionShowBasemap2Layer.triggered.connect(self.showBasemap2Layer)
        self.actionShowBasemap3Layer.triggered.connect(self.showBasemap3Layer)
        self.actionZoomIn.triggered.connect(self.zoomIn)
        self.actionZoomOut.triggered.connect(self.zoomOut)
        self.actionPan.triggered.connect(self.setPanMode)
        self.actionOH.triggered.connect(self.setOriginalHouse)
        self.actionG.triggered.connect(self.setGroundMap)
        self.actionF.triggered.connect(self.setFirstMap)
        self.actionE1.triggered.connect(self.setExt1Map)
        self.actionE2.triggered.connect(self.setExt2Map)
        self.actionA.triggered.connect(self.setAtticMap)
        self.actionL.triggered.connect(self.setLineMap)
        self.actionAP.triggered.connect(self.setAllPlanningMap)
        self.actionEditLine.triggered.connect(self.setEditLineMode)
        self.actionEditPolygon.triggered.connect(self.setEditPolygonMode)
        self.actionAdd.triggered.connect(self.addFeature)
        self.actionEdit.triggered.connect(self.editFeature)
        self.actionDelete.triggered.connect(self.deleteFeature)
        self.actionSplitPoly.triggered.connect(self.splitPoly)
        self.actionManualTileExt.triggered.connect(self.manualTileExt)
        self.actionManualTileLoft.triggered.connect(self.manualTileLoft)
        self.actionGetInfo.triggered.connect(self.getInfo)
        self.actionManualMeasures.triggered.connect(self.manualMeasures)
        self.actionManualRules.triggered.connect(self.manualRules)
        self.actionManualExtensions.triggered.connect(self.manualExtensions)
        self.actionManualLoft.triggered.connect(self.manualLoft)
        self.actionManualCombo.triggered.connect(self.manualCombo)
        self.actionReport.triggered.connect(self.report)
        self.actionSelectFeature.changed.connect(self.onSelectFeaturesChanged)
        self.actionImportFeatures.triggered.connect(self.onImportFeaturesTriggered)
        self.actionGraph.triggered.connect(self.graphFunction)
        self.actionGraphPP.triggered.connect(self.graphFunctionPP)
        self.actionGraphAP.triggered.connect(self.graphFunctionAP)
        self.actionShowScrape.triggered.connect(self.displayScrape)
        self.actionPlannerNotes.triggered.connect(self.plannerNotes)

        self.selectFeatureTool = None

        self.mapCanvas = QgsMapCanvas()
        self.mapCanvas.setCanvasColor(Qt.white)
        self.mapCanvas.show()
        self.mapCanvas.scaleChanged.connect(self.onMapCanvasScaleChanged)
        self.mapCanvas.mapToolSet.connect(self.onCanvasMapToolSet)

        self.actionShowBasemap1Layer.setChecked(True)

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.mapCanvas)

        self.centralWidget.setLayout(layout)

        self.panTool = PanTool(self.mapCanvas)
        self.panTool.setAction(self.actionPan)

        self.layer_num = 2
        self.actionG.setChecked(True)
        self.installEventFilter(self)

        self.webscraping_file = '/MyComputer/Dropbox/Lanu/houses/houses_webscraping_test.xlsx'
        self.house_folder_rel = '/MyComputer/Dropbox/Lanu/houses'
        self.testingCRUD_file = '/MyComputer/Dropbox/Lanu/extensions_rules.xlsx'

        self.datum = {
            'id': '',
            'address': '',
            'link': '',
            'description': 'Semi Detached',
            'date': '',
            'price': 0,
            'type house': 'Semi Detached',
            'conservation_area': 'No',
            'green_belt': 'No',
            'image_url': '',
            'bedrooms': 3,
            'bathrooms': 1,
            'reception': 2,
            'area': 0,
            'Borough': 0,
            'PPSQM': 0,
            'BPSQM': 0,
            'summary_report_bool': False,
            'add_value_bool': False,
            'n_rooms': 5,
            'EPC': '',
            'AGE': ''
        }

        QTimer.singleShot(0, lambda: self.showInitialDialog())

    def new_project_pickle_files(self):
        fileName = os.path.join(self.dropbox_foldername, '/measures.pickle')
        if not os.path.exists(fileName):
            a = {'HW': -100, 'HD': -100, 'SW': -100, 'SD': -100, 'SDR': -100, 'RDGL': -100,
                 'GB1': -100, 'GFB1': -100, 'GSE1': -100, 'GFN1': -100, 'FFN1': -100, 'GFNF1': -100, 'FFNF1': -100,
                 'GB2': -100, 'GFB2': -100, 'GSE2': -100, 'GFN2': -100, 'FFN2': -100, 'GFNF2': -100, 'FFNF2': -100,
                 'H': 3, 'H1': 3, 'HL': 3, 'HO3': 3, 'HO4': 3, 'HO5': 3, 'HO6': 3, 'HO19': 3, 'HO20': 3,
                 'RHO3': 3, 'RHO4': 3, 'RHO5': 3, 'RHO6': 3, 'RHO19': 3, 'RHO20': 3}
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        fileName = os.path.join(self.dropbox_foldername, '/extensions.pickle')
        if not os.path.exists(fileName):
            a = {'E': [], 'E_PP': [], 'E_AP': [], 'Ext_list': [], 'Ext_PP_list': [], 'Ext_AP_list': [],
                 'Push_to_combo': False}
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        fileName = os.path.join(self.dropbox_foldername, '/loft.pickle')
        if not os.path.exists(fileName):
            a = {'EL': [], 'EL_PP': [], 'EL_AP': [], 'Ext_loft_list': [], 'Push_to_combo': False}
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        fileName = os.path.join(self.dropbox_foldername, '/combos.pickle')
        if not os.path.exists(fileName):
            a = {'GF': False, 'GF_list': '', 'FF': False, 'FF_list': '', 'AT': False, 'AT_list': '',
                 'GF_PP': False, 'GF_list_PP': '', 'FF_PP': False, 'FF_list_PP': '', 'AT_PP': False, 'AT_list_PP': '',
                 'GF_AP': False, 'GF_list_AP': '', 'FF_AP': False, 'FF_list_AP': '', 'AT_AP': False, 'AT_list_AP': ''}
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        fileName = os.path.join(self.dropbox_foldername, '/planner_notes.pickle')
        if not os.path.exists(fileName):
            a = {'planner_notes': ''}
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

    def showInitialDialog(self):
        """Initial projects loader/selector dialog """
        if not self.openProjectDialog():
            self.close()

    def openProjectDialog(self):
        """Open projects loader/selector dialog, returns False if Cancel button was pressed"""
        if self.checkSaveProject():
            if DEBUG:
                address = '67 Lynmouth Drive, Ruislip, HA4 9DT, UK'
            else:
                address = ''

            dialog = ProjectLoadDialog(parent=self, address=address, house_folder_rel=self.house_folder_rel)
            if dialog.exec() == QDialog.Accepted:
                print('*********')
                self.loadProject(dialog.dropbox_foldername, dialog.center)
                return True
            else:
                return False

        return True

    def showManualMeasures(self):
        dialog = ShowManualMeasures(parent=self)
        if dialog.exec() == QDialog.Accepted:
            pass

    def showExtensionImage(self):
        dialog = ShowExtensions(parent=self)
        if dialog.exec() == QDialog.Accepted:
            pass

    def showLanesImage(self):
        dialog = ShowLanesImage(parent=self)
        if dialog.exec() == QDialog.Accepted:
            pass

    def showLoftImage(self):
        dialog = ShowLoftImage(parent=self)
        if dialog.exec() == QDialog.Accepted:
            pass

    def check_spreadsheet(self):
        # set house variables to default values
        self.set_default_values()
        # get house variables from web scraping spread sheet if present
        house_list = self.get_data_from_workbook()
        # if house in workbook
        if house_list != []:
            self.assign_house_variables_by_list(house_list)
        # if house is not in workbook then use the web scraper
        else:
            print('Webscraping')
            process = self.webscrape()
            self.datum = process.datum
            if process == False:
                self.scrapeNotDone = True
            else:
                self.scrapeNotDone = False
                self.set_variables_from_datum()

        # open dialog box to manually edit web scraping results / house variables
        check, checkaccepted = self.checkwebscrape(self.datum)

        self.datum = check
        process = ProcessSingleHouse(self.address_list, self.link_list, self.id_list)
        # write results to web scraping spread sheet
        process.write_to_spreadsheet_alt(self.datum)

        # update house variables
        self.set_variables_from_datum()

        print('id_list: ', self.id_list)
        print('address_list: ', self.address_list)
        print('link_list: ', self.link_list)
        print('description_list: ', self.description_list)
        print('date_list: ', self.date_list)
        print('price_list: ', self.price_list)
        print('type_list: ', self.type_list)
        print('conservation_area_list: ', self.conservation_area_list)
        print('green_belt_list: ', self.green_belt_list)
        print('image_url_list: ', self.image_url_list)
        print('bedrooms_list: ', self.bedrooms_list)
        print('bathrooms_list: ', self.bathrooms_list)
        print('reception_list: ', self.reception_list)
        print('area_list: ', self.area_list)
        print('Borough_list: ', self.Borough_list)
        print('PPSQM_list: ', self.PPSQM_list)
        print('BPSQM_list: ', self.BPSQM_list)
        print('summary_report_list: ', self.summary_report_list)
        print('add_value_list: ', self.add_value_list)
        print('n_rooms: ', self.n_rooms_list)
        print('EPC: ', self.EPC_list)
        print('AGE: ', self.AGE_list)

    def set_default_values(self):
        reobj = re.compile(r'(\b[A-Z]{1,2}[0-9][A-Z0-9]? [0-9][ABD-HJLNP-UW-Z]{2}\b)')
        matchobj = reobj.search(self.address)
        if matchobj:
            self.postcode = matchobj.group(1)
            self.postcode.replace("", "_")
        else:
            self.postcode = "------"

        self.address_temp = self.address.replace(',', '')
        self.address_temp = self.address_temp.replace('.', '')
        self.address_temp = self.address_temp.split()
        try:
            self.house_id = self.address_temp[0] + '_' + self.postcode
            self.house_id = self.house_id.replace(" ", "_")
        except:
            self.house_id = ''

        self.id_list = self.house_id
        self.address_list = self.address
        self.link_list = ''
        self.description_list = '3 Bed Detached'
        now = datetime.datetime.now()
        uf = UsefulFunctions()
        day = uf.string_day(str(now.day))
        month = uf.string_month(str(now.month))
        year = str(now.year)
        self.date_list = day + ' ' + month + ' ' + year
        self.price_list = 0.0
        self.type_list = 'Detached'
        self.conservation_area_list = 'No'
        self.green_belt_list = 'No'
        self.image_url_list = ''
        self.bedrooms_list = 3
        self.bathrooms_list = 1
        self.reception_list = 2
        self.area_list = 0
        self.Borough_list = 'None'
        self.PPSQM_list = 0
        self.BPSQM_list = 0
        self.summary_report_list = False
        self.add_value_list = False
        self.n_rooms_list = 5
        self.EPC_list = ""
        self.AGE_list = ""

    def set_variables_from_datum(self):
        self.id_list = self.datum['id']
        self.address_list = self.datum['address']
        self.link_list = self.datum['link']
        self.description_list = self.datum['description']
        self.date_list = self.datum['date']
        self.price_list = self.datum['price']
        self.type_list = self.datum['type house']
        self.conservation_area_list = self.datum['conservation_area']
        self.green_belt_list = self.datum['green_belt']
        self.image_url_list = self.datum['image_url']
        self.bedrooms_list = self.datum['bedrooms']
        self.bathrooms_list = self.datum['bathrooms']
        self.reception_list = self.datum['reception']
        self.area_list = self.datum['area']
        self.Borough_list = self.datum['Borough']
        self.PPSQM_list = self.datum['PPSQM']
        self.BPSQM_list = self.datum['BPSQM']
        self.summary_report_list = self.datum['summary_report_bool']
        self.add_value_list = self.datum['add_value_bool']
        self.n_rooms_list = self.datum['n_rooms']
        self.EPC_list = self.datum['EPC']
        self.AGE_list = self.datum['AGE']

    def get_data_from_workbook(self):
        wb = openpyxl.load_workbook(self.webscraping_file)
        ws = wb['Sheet1']

        idlist = []
        for row in ws['A']:
            idlist.append(row.value)

        if self.id_list in idlist:
            ind = idlist.index(self.id_list)
            house_list_temp = list(ws.rows)[ind]
            house_list = [house_list_temp[x].value for x in range(len(house_list_temp))]
        else:
            house_list = []
        wb.close()
        # Close workbook

        return house_list

    def assign_house_variables_by_list(self, house_list):
        self.address_list = house_list[1]
        self.link_list = house_list[2]
        self.description_list = house_list[3]
        self.date_list = house_list[4]
        self.price_list = house_list[5]
        self.type_list = house_list[6]
        self.conservation_area_list = house_list[7]
        self.green_belt_list = house_list[8]
        self.image_url_list = house_list[9]
        self.bedrooms_list = house_list[10]
        self.bathrooms_list = house_list[11]
        self.reception_list = house_list[12]
        self.area_list = house_list[13]
        self.Borough_list = house_list[14]
        self.PPSQM_list = house_list[15]
        self.BPSQM_list = house_list[16]
        self.summary_report_list = house_list[17]
        self.add_value_list = house_list[18]
        self.n_rooms_list = house_list[19]
        self.EPC_list = house_list[20]
        self.AGE_list = house_list[21]

        self.datum = {
            'id': self.id_list,
            'address': self.address_list,
            'link': self.link_list,
            'description': self.description_list,
            'date': self.date_list,
            'price': self.price_list,
            'type house': self.type_list,
            'conservation_area': self.conservation_area_list,
            'green_belt': self.green_belt_list,
            'image_url': self.image_url_list,
            'bedrooms': self.bedrooms_list,
            'bathrooms': self.bathrooms_list,
            'reception': self.reception_list,
            'area': self.area_list,
            'Borough': self.Borough_list,
            'PPSQM': self.PPSQM_list,
            'BPSQM': self.BPSQM_list,
            'summary_report_bool': self.summary_report_list,
            'add_value_bool': self.add_value_list,
            'n_rooms': self.n_rooms_list,
            'EPC': self.EPC_list,
            'AGE': self.AGE_list
        }

    def compareAddres(self, cmp_address, postcode, number):

        cmp_number = None
        cmp_postcode = None

        reobj = re.compile(r'(\b[A-Z]{1,2}[0-9][A-Z0-9]? [0-9][ABD-HJLNP-UW-Z]{2}\b)')
        matchobj = reobj.search(cmp_address)
        if matchobj:
            cmp_postcode = matchobj.group(1)
            cmp_postcode.replace("", "_")

        cmp_address = cmp_address.split()
        cmp_number = cmp_address[:2]

        if cmp_postcode == postcode and number == cmp_number:
            correct = True
        else:
            correct = False

        return correct

    def loadProject(self, dropbox_foldername, center):

        """Loads a project either an existing or a new one"""
        self.dropbox_foldername = dropbox_foldername
        self.center = center

        self.housename = "_" + os.path.basename(self.dropbox_foldername)
        # Update address from geocoder
        self.address = self.housename.replace("_", " ").replace('-', ' ')
        self.address = self.address.lstrip()
        self.setWindowTitle(self.address)

        # List of QgsHighlight
        self.highlights = []
        self.highlighted_features = []

        self.layer_id = ''
        self.map_id = ''
        self.housetype = ''

        # Manual Measures and Extensions
        b = {'HW': -100, 'HD': -100, 'SW': -100, 'SD': -100, 'SDR': -100, 'RDGL': -100,
             'GB1': -100, 'GFB1': -100, 'GSE1': -100, 'GFN1': -100, 'FFN1': -100, 'GFNF1': -100, 'FFNF1': -100,
             'GB2': -100, 'GFB2': -100, 'GSE2': -100, 'GFN2': -100, 'FFN2': -100, 'GFNF2': -100, 'FFNF2': -100,
             'H': 3, 'H1': 3, 'HL': 3, 'HO3': 3, 'HO4': 3, 'HO5': 3, 'HO6': 3, 'HO19': 3, 'HO20': 3,
             'RHO3': 3, 'RHO4': 3, 'RHO5': 3, 'RHO6': 3, 'RHO19': 3, 'RHO20': 3}

        fileName = self.dropbox_foldername + '/measures.pickle'
        if not os.path.exists(fileName):
            with open(fileName, 'wb') as handle:
                pickle.dump(b, handle, protocol=pickle.HIGHEST_PROTOCOL)

        ## Extension Stuff ##
        self.Ext_list = ["Fix PD", "Ext 1", "Ext 2", "Ext 3", "Ext 4", "Ext 5", "Ext 6", "Ext 19", "Ext 20",
                         "Ext 7", "Ext 8", "Ext 9", "Ext 10", "Ext 11", "Ext 12", "Ext 18", "Ext 21", "Ext 22",
                         "Ext 23", "Ext 13", "Ext 3a", "Ext 4a", "Ext 5a", "Ext 6a", "Ext 19a", "Ext 20a", '',
                         "Push To Combos"]
        self.Ext_PP_list = self.Ext_list[:]
        self.Ext_PP_list[0] = "Fix PP"
        self.Ext_AP_list = self.Ext_list[:]
        self.Ext_AP_list[0] = "Fix AP"

        self.E = [False] * len(self.Ext_list)
        self.E_PP = [False] * len(self.Ext_list)
        self.E_AP = [False] * len(self.Ext_list)
        self.E[0], self.E_PP[0], self.E_AP[0] = True, True, True
        a = {'E': self.E, 'E_PP': self.E_PP, 'E_AP': self.E_AP, 'Ext_list': self.Ext_list,
             'Ext_PP_list': self.Ext_PP_list,
             'Ext_AP_list': self.Ext_AP_list, 'Push_to_combo': False}
        fileName = self.dropbox_foldername + '/extensions.pickle'
        if not os.path.exists(fileName):
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        ## Loft Stuff ##
        self.Ext_loft_list = ["Fix", "Hip to Gable [1]", "Side Dormer [2]", "Half Width Rear Dormer (Open) [3]",
                              "Full Width Rear Dormer [4]", "Half Width rear Dormer (Closed) [5]",
                              "Rear Dormer (centred) [6]",
                              "Hip to Gable(opposite) [7]", "Outrigger Dormer [8]", "", "Push To Combos"]
        self.EL = [False] * len(self.Ext_loft_list)
        self.EL_PP = [False] * len(self.Ext_loft_list)
        self.EL_AP = [False] * len(self.Ext_loft_list)
        self.EL[0], self.EL_PP[0], self.EL_AP[0] = True, True, True

        a = {'EL': self.EL, 'EL_PP': self.EL_PP, 'EL_AP': self.EL_AP, 'Ext_loft_list': self.Ext_loft_list,
             'Push_to_combo': False}
        fileName = self.dropbox_foldername + '/loft.pickle'
        if not os.path.exists(fileName):
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        ## Combo Stuff ##
        self.gf_manual, self.ff_manual, self.at_manual = False, False, False
        self.gf_manual_combo_list, self.ff_manual_combo_list, self.at_manual_combo_list = '', '', ''

        self.gf_manual, self.ff_manual, self.at_manual = False, False, False
        self.gf_manual_combo_list, self.ff_manual_combo_list, self.at_manual_combo_list = '', '', ''

        self.gf_PP_manual, self.ff_PP_manual, self.at_PP_manual = False, False, False
        self.gf_PP_manual_combo_list, self.ff_PP_manual_combo_list, self.at_PP_manual_combo_list = '', '', ''

        self.gf_AP_manual, self.ff_AP_manual, self.at_AP_manual = False, False, False
        self.gf_AP_manual_combo_list, self.ff_AP_manual_combo_list, self.at_AP_manual_combo_list = '', '', ''

        a = {'GF': self.gf_manual, 'GF_list': self.gf_manual_combo_list, 'FF': self.ff_manual,
             'FF_list': self.ff_manual_combo_list,
             'AT': self.at_manual, 'AT_list': self.at_manual_combo_list,
             'GF_PP': self.gf_PP_manual, 'GF_list_PP': self.gf_PP_manual_combo_list, 'FF_PP': self.ff_PP_manual,
             'FF_list_PP': self.ff_PP_manual_combo_list,
             'AT_PP': self.at_PP_manual, 'AT_list_PP': self.at_PP_manual_combo_list,
             'GF_AP': self.gf_AP_manual, 'GF_list_AP': self.gf_AP_manual_combo_list, 'FF_AP': self.ff_AP_manual,
             'FF_list_AP': self.ff_AP_manual_combo_list,
             'AT_AP': self.at_AP_manual, 'AT_list_AP': self.at_AP_manual_combo_list}

        fileName = self.dropbox_foldername + '/combos.pickle'
        if not os.path.exists(fileName):
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        a = {'planner_notes': ''}
        fileName = self.dropbox_foldername + '/planner_notes.pickle'
        if not os.path.exists(fileName):
            with open(fileName, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        ## Check spreadsheet ##
        # try:
        self.check_spreadsheet()
        # except:
        #     print('Some problem with spreadsheet')

        ## Write the small excel file

        fileName = self.dropbox_foldername + '/' + os.path.basename(self.dropbox_foldername) + '.xlsx'
        # if not os.path.exists(fileName):
        workbook = xlsxwriter.Workbook(fileName)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'description')
        worksheet.write(0, 1, 'name')
        worksheet.write(0, 2, 'address')
        worksheet.write(0, 3, 'date')
        worksheet.write(0, 4, 'price')
        worksheet.write(0, 5, 'phone')
        worksheet.write(0, 6, 'link')
        worksheet.write(0, 7, 'link_house')
        worksheet.write(0, 8, 'type')
        worksheet.write(0, 9, 'conservation_area')
        worksheet.write(0, 10, 'green_belt')
        worksheet.write(0, 11, 'Borough')
        worksheet.write(0, 12, 'Image_url')
        worksheet.write(0, 13, 'PPSQM')
        worksheet.write(0, 14, 'BPSQM')
        worksheet.write(0, 15, 'summary_report_bool')
        worksheet.write(0, 16, 'add_value_bool')
        worksheet.write(0, 17, 'n_rooms')
        worksheet.write(0, 18, 'EPC')
        worksheet.write(0, 19, 'AGE')

        worksheet.write(1, 0, self.description_list)
        worksheet.write(1, 1, os.path.basename(self.dropbox_foldername))
        worksheet.write(1, 2, self.address)
        worksheet.write_string(1, 3, self.date_list)
        worksheet.write(1, 4, self.price_list)
        worksheet.write(1, 5, "")
        worksheet.write(1, 6, self.link_list)
        worksheet.write(1, 7, "")
        worksheet.write(1, 8, self.type_list)
        worksheet.write(1, 9, self.conservation_area_list)
        worksheet.write(1, 10, self.green_belt_list)
        worksheet.write(1, 11, self.Borough_list)
        worksheet.write(1, 12, self.image_url_list)
        worksheet.write(1, 13, self.PPSQM_list)
        worksheet.write(1, 14, self.BPSQM_list)
        worksheet.write(1, 15, self.summary_report_list)
        worksheet.write(1, 16, self.add_value_list)
        worksheet.write(1, 17, self.n_rooms_list)
        worksheet.write(1, 18, self.EPC_list)
        worksheet.write(1, 19, self.AGE_list)

        workbook.close()

        ## Getting png
        fileName = self.dropbox_foldername + "/property.png"
        if not os.path.exists(fileName):
            try:
                if len(self.image_url_list) > 0:
                    if self.image_url_list[-2:] == ':p':
                        self.image_url_list = self.image_url_list[:-2]
                    img_data = requests.get(self.image_url_list).content
                    file = open(fileName, "wb")
                    file.write(img_data)
                    self.resize(self.sizeHint())
            except:
                cmd = 'cp ../../Dropbox/Lanu/property.png ' + fileName
                os.system(cmd)

        ## data file
        a = {'type_list': self.type_list}

        fileName = self.dropbox_foldername + '/data.pickle'
        with open(fileName, 'wb') as handle:
            pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

        for i in range(1, NUMBER_OF_SQLITE_HOUSE_LAYERS + 1):
            self.setupDatabase(self.housename + "_%s" % i)

        ## final methods
        self.loadMap()
        self.setupMapTools()

        self.setPanMode()
        self.adjustActions()

    def createOsLayers(self):
        """Creates the OS layers with base map and buildings"""

        self.osBaseLayer = QgsRasterLayer(
            'contextualWMSLegend=0&crs=EPSG:27700&dpiMode=7&featureCount=10&format=image/png&layers=Light_27700&styles=default&tileMatrixSet=EPSG:27700&url=https://api.os.uk/maps/raster/v1/wmts?key%3D' + OS_API_KEY,
            'OS Base Map Light 27700', 'wms')

        self.osBuildingsLayer = QgsVectorLayer(
            " pagingEnabled='true' preferCoordinatesForWfsT11='false' restrictToRequestBBOX='1' srsname='EPSG:27700' typename='osfeatures:Topography_TopographicArea' url='https://api.os.uk/features/v1/wfs?key=%s' version='auto'" % OS_API_KEY,
            'OS Topography Polygons', 'WFS')

        self.osBuildingsLayer.loadNamedStyle(os.path.join(
            os.path.dirname(os.path.realpath(__file__)), 'assets', 'osfeatures_topographic_area.qml'))

        QgsProject.instance().addMapLayers(
            [self.osBaseLayer, self.osBuildingsLayer])

    def addSatelliteBaseMap(self):
        """Adds a satellite basemap layer"""

        self.satelliteBaseMap = QgsRasterLayer(
            'type=xyz&url=%s&zmax=18&zmin=0' % quote_plus(SATELLITE_BASEMAP), 'Satellite', 'wms')

        QgsProject.instance().addMapLayers([self.satelliteBaseMap])

    def onSelectFeaturesChanged(self):
        """Signal handler for select feature action"""

        if self.actionSelectFeature.isChecked():
            self.selectFeatureTool = SelectFeatureTool(
                self.mapCanvas, self.osBuildingsLayer, self.onFeatureSelected)
            self.mapCanvas.setMapTool(self.selectFeatureTool)

    def onFeatureSelected(self, feature):
        """Feature selected callback

        :param feature: selected feature
        :type feature: QgsFeature
        """

        removed = False
        highlights = []
        highlighted_features = []

        for highlighted_feature in self.highlighted_features:
            if highlighted_feature.geometry().asWkt() == feature.geometry().asWkt():
                removed = True
            else:
                highlighted_features.append(highlighted_feature)

        if not removed:
            highlighted_features.append(feature)

        self.highlighted_features = highlighted_features

        for h in self.highlights:
            self.mapCanvas.scene().removeItem(h)

        for highlighted_feature in self.highlighted_features:
            highlight = QgsHighlight(self.mapCanvas, highlighted_feature.geometry(), self.osBuildingsLayer)
            highlight.setWidth(4)
            highlight.setColor(QColor(255, 255, 0))
            highlight.setFillColor(QColor(255, 255, 255, 0))
            highlights.append(highlight)

        self.highlights = highlights

        logging.debug('Selected feature %s: %s' % ('removed' if removed else 'added', feature.attributes()[0]))

    def onImportFeaturesTriggered(self):
        """Signal handler for import selected features action.

        All selected features will imported into the "active" layer, which is
        determined by the layer_num value.

        The application must be in editing mode and the active layer must be editable.

        """

        self.active_layer = getattr(self, 'houseLayer%s' % self.layer_num)

        if not self.editing or not self.active_layer.isEditable():
            QMessageBox.warning(None, "Error importing features",
                                "There is no current edited layer")
        else:

            new_features = []

            for feature in self.highlighted_features:
                new_feature = QgsFeature(self.active_layer.fields())
                geometry = feature.geometry()
                if QgsWkbTypes.singleType(geometry.wkbType()) == QgsWkbTypes.LineString and QgsWkbTypes.singleType(
                        self.active_layer.wkbType()) == QgsWkbTypes.Polygon:
                    geometry = geometry.convexHull()
                elif QgsWkbTypes.singleType(geometry.wkbType()) == QgsWkbTypes.Polygon and QgsWkbTypes.singleType(
                        self.active_layer.wkbType()) == QgsWkbTypes.LineString:
                    p = next(geometry.parts())
                    c = p.toCurveType()
                    geometry = c.exteriorRing().clone()

                points = geometry.points()
                x = []
                y = []
                for i in range(len(points)):
                    x.append(points[i].x())
                    y.append(points[i].y())

                # If dealing with ahouse remove colinear vertices
                if self.layer_num in [1, 2, 3, 7]:
                    gt = geometrytools()
                    gt.tol = 0.1
                    x, y = gt.remove_close_points(x, y)
                    x, y = gt.remove_colinear_points(x, y)
                    x.append(x[0])
                    y.append(y[0])

                points = []
                for i in range(len(x)):
                    points.append(QgsPointXY(x[i], y[i]))
                geometry_ = QgsGeometry.fromPolylineXY(points)
                new_feature.setGeometry(geometry_)
                new_feature = self.saveAttributes_import(new_feature, self.layer_num)
                new_features.append(new_feature)

            if self.layer_num in [4, 5, 6]:
                new_features = self.combine_features(new_features)

            self.active_layer.addFeatures(new_features)

            self.modified = True

            # Clear all highlights and selections
            for h in self.highlights:
                self.mapCanvas.scene().removeItem(h)

            self.highlights = []
            self.highlighted_features = []
            self.mapCanvas.refresh()

    def combine_features(self, features):
        attributes = features[0].attributes()

        polygons = []
        for feature in features:
            geometry = feature.geometry()
            points = geometry.asPolyline()
            x = []
            y = []
            for i in range(len(points)):
                x.append(points[i].x())
                y.append(points[i].y())
            polygons.append(Polygon(zip(x, y)))

        u = cascaded_union(polygons)
        x, y = u.exterior.coords.xy
        x_new, y_new = list(x), list(y)

        points = []
        for j in range(len(x_new)):
            points.append(QgsPointXY(x_new[j], y_new[j]))
        new_geometry = QgsGeometry.fromPolylineXY(points)

        new_feature = QgsFeature(self.active_layer.fields())
        new_feature.setGeometry(new_geometry)

        new_feature.setAttributes(attributes)
        features = [new_feature]

        return features

    def num_features(self):
        n_features = 0
        max_features = 0
        feature_ids = []
        for feature in self.active_layer.getFeatures():
            n_features += 1
            attributes = feature.attributes()
            feature_ids.append(attributes[0])
            max_features = max(feature_ids)
        return n_features, max_features

    def saveAttributes_import(self, feature, layer_num):

        if self.type_list == 'Semi-Detached':
            house_type_attr = HOUSE_TYPE_SD
        elif self.type_list == 'Terraced':
            house_type_attr = HOUSE_TYPE_T
        elif self.type_list == 'Detached':
            house_type_attr = HOUSE_TYPE_D
        else:
            house_type_attr = HOUSE_TYPE_SD

        floor_attr = TW_S

        if self.curEditedLayer == self.lineLayer and layer_num in [5]:
            part_attr = HOUSE_FRONT
        elif layer_num in [1, 2, 3, 7, 8]:
            part_attr = GROUND
        elif layer_num in [4, 5, 6]:
            part_attr = SITE
        else:
            part_attr = NONE

        if layer_num in [1, 4]:
            main_house_attr = LEFT_HOUSE
        if layer_num in [2, 5, 7, 8]:
            main_house_attr = MAIN_HOUSE
        if layer_num in [3, 6]:
            main_house_attr = RIGHT_HOUSE

        if layer_num in [7]:
            pd_attr = PD_N
            pp_attr = PP_Y
            ap_attr = AP_N
        elif layer_num in [8]:
            pd_attr = PD_N
            pp_attr = PP_N
            ap_attr = AP_Y
        elif layer_num in [4, 5, 6]:
            pd_attr = PD_Y
            pp_attr = PP_Y
            ap_attr = AP_Y
        else:
            pd_attr = PD_Y
            pp_attr = PP_N
            ap_attr = AP_N

        main_house_ext_attr = ENone
        main_house_desc_attr = DNone
        color_attr = NONE_COLOR
        manual_attr = MAN_N
        exist_attr = EXIST_N
        cnv_attr = CONV_N
        sab1_attr = SAB1_N
        nn1_attr = NN1_N
        basement_attr = BA_N
        or_attr = OR_Y
        gr_attr = GR_N
        hr_attr = HR_N
        bgr_attr = BGR_N
        crf_attr = CRF_N
        dg_attr = DG_N
        ne_attr = NE_N
        hnd_attr = HND_N
        ucr_attr = UCR_N
        otr_attr = OTR_N
        et_attr = ET_N
        rw_attr = RW_N
        fw_attr = FW_N

        n_features, max_features = self.num_features()
        feature.setAttribute("ogc_fid", max_features + 1)
        feature.setAttribute("house_type", house_type_attr)
        feature.setAttribute("floor", floor_attr)
        feature.setAttribute("part", part_attr)
        feature.setAttribute("main_house", main_house_attr)
        feature.setAttribute("ext", main_house_ext_attr)
        feature.setAttribute("desc", main_house_desc_attr)
        feature.setAttribute("color", color_attr)
        feature.setAttribute("manual", manual_attr)
        feature.setAttribute("exist", exist_attr)
        feature.setAttribute("cnv", cnv_attr)
        feature.setAttribute("sab1", sab1_attr)
        feature.setAttribute("nn1", nn1_attr)
        feature.setAttribute("orig", or_attr)
        feature.setAttribute("gr", gr_attr)
        feature.setAttribute("hr", hr_attr)
        feature.setAttribute("bgr", bgr_attr)
        feature.setAttribute("crf", crf_attr)
        feature.setAttribute("dg", dg_attr)
        feature.setAttribute("ne", ne_attr)
        feature.setAttribute("hnd", hnd_attr)
        feature.setAttribute("ucr", ucr_attr)
        feature.setAttribute("otr", otr_attr)
        feature.setAttribute("et", et_attr)
        feature.setAttribute("pd", pd_attr)
        feature.setAttribute("pp", pp_attr)
        feature.setAttribute("ap", ap_attr)
        feature.setAttribute("rw", rw_attr)
        feature.setAttribute("fw", fw_attr)
        feature.setAttribute("height", '3.0')
        feature.setAttribute("basement", basement_attr)

        return feature

    def onMapCanvasScaleChanged(self, scaleDenominator):
        """Signal handler for map canvas scale/center changed"""

        logger.debug("Scale changed to 1:%s" % scaleDenominator)
        self.updateStatusBar()
        self.showVisibleMapLayers()

    def onCanvasMapToolSet(self, newTool, oldTool):
        """Signal handler: Reset select features tools state"""

        if newTool != self.selectFeatureTool:
            self.actionSelectFeature.setChecked(False)

    def updateStatusBar(self, distance=-1):
        """Signal handler: Updates the status bar

        :param distance: optional distance to show, defaults to -1
        :type distance: int, optional
        """

        self.statusBar().showMessage(self.statusBarText(distance))

    def statusBarText(self, distance=-1):
        """Status bar text

        :param distance: optional distance to show, defaults to -1
        :type distance: int, optional
        """
        text = "Scale: 1:%s" % int(self.mapCanvas.scale())

        if distance >= 0:
            text += " - distance: %0.2f m" % distance

        return text

    def eventFilter(self, source, event):

        if event.type() == QEvent.MouseMove:
            self.updateStatusBar()
        return QMainWindow.eventFilter(self, source, event)

    def setupDatabase(self, filename):
        """Create sqlite layer

        :param filename: filename
        :type filename: str
        """

        str1 = filename + ".sqlite"
        dbName = os.path.join(self.dropbox_foldername, str1)
        print(dbName)
        if not os.path.exists(dbName):

            fields = QgsFields()
            fields.append(QgsField("house_type", QVariant.String))
            fields.append(QgsField("floor", QVariant.String))
            fields.append(QgsField("part", QVariant.String))
            fields.append(QgsField("main_house", QVariant.String))
            fields.append(QgsField("ext", QVariant.String))
            fields.append(QgsField("desc", QVariant.String))
            fields.append(QgsField("color", QVariant.String))
            fields.append(QgsField("manual", QVariant.String))
            fields.append(QgsField("exist", QVariant.String))
            fields.append(QgsField("cnv", QVariant.String))
            fields.append(QgsField("sab1", QVariant.String))
            fields.append(QgsField("nn1", QVariant.String))
            fields.append(QgsField("orig", QVariant.String))
            fields.append(QgsField("gr", QVariant.String))
            fields.append(QgsField("hr", QVariant.String))
            fields.append(QgsField("bgr", QVariant.String))
            fields.append(QgsField("crf", QVariant.String))
            fields.append(QgsField("dg", QVariant.String))
            fields.append(QgsField("ne", QVariant.String))
            fields.append(QgsField("hnd", QVariant.String))
            fields.append(QgsField("ucr", QVariant.String))
            fields.append(QgsField("otr", QVariant.String))
            fields.append(QgsField("et", QVariant.String))
            fields.append(QgsField("pd", QVariant.String))
            fields.append(QgsField("pp", QVariant.String))
            fields.append(QgsField("ap", QVariant.String))
            fields.append(QgsField("rw", QVariant.String))
            fields.append(QgsField("fw", QVariant.String))
            fields.append(QgsField("mincost", QVariant.String))
            fields.append(QgsField("maxcost", QVariant.String))
            fields.append(QgsField("minarea", QVariant.String))
            fields.append(QgsField("maxarea", QVariant.String))
            fields.append(QgsField("kitchen", QVariant.String))
            fields.append(QgsField("living", QVariant.String))
            fields.append(QgsField("storage", QVariant.String))
            fields.append(QgsField("bedroom", QVariant.String))
            fields.append(QgsField("bathroom", QVariant.String))
            fields.append(QgsField("keepEE", QVariant.String))
            fields.append(QgsField("height", QVariant.String))
            fields.append(QgsField("basement", QVariant.String))

            crs = QgsCoordinateReferenceSystem('EPSG:27700')
            writer = QgsVectorFileWriter(dbName, 'utf-8', fields,
                                         QgsWkbTypes.LineString,
                                         crs, 'SQLite',
                                         ["SPATIALITE=YES"])

            if writer.hasError() != QgsVectorFileWriter.NoError:
                print("Error creating measures database!")

            del writer

    def loadMap(self):
        """Load the map layers from the sqlite files"""

        cur_dir = os.path.dirname(os.path.realpath(__file__))

        # Add base layers
        QgsProject.instance().read(os.path.join(cur_dir, 'assets', 'lanu_base_project.qgs'))

        # Add house layers
        color_list = ['', 'blue', 'red', 'orange', 'magenta', 'mediumturquoise', 'yellow', 'lime', 'black', 'grey']
        for idx in range(1, NUMBER_OF_SQLITE_HOUSE_LAYERS + 1):

            filename = self.housename + "_%s" % idx
            uri = QgsDataSourceUri()
            uri.setDatabase(os.path.join(self.dropbox_foldername, filename + ".sqlite"))
            uri.setDataSource('', filename, 'GEOMETRY')
            layer_name = 'houseLayer%s' % idx
            setattr(self, layer_name, QgsVectorLayer(
                uri.uri(), filename, "spatialite"))
            layer = getattr(self, layer_name)
            s = QgsLineSymbol.createSimple({"color": color_list[8], "width": "1"})
            r = QgsSingleSymbolRenderer(s)
            layer.setRenderer(r)
            symbol = layer.renderer().symbol()
            symbol.setWidth(1)
            if color_list[idx] == 'blue':
                symbol.setColor(QColor(0, 0, 255, 255))
            elif color_list[idx] == 'red':
                symbol.setColor(QColor(255, 0, 0, 255))
            elif color_list[idx] == 'orange':
                symbol.setColor(QColor(255, 165, 0, 255))
            elif color_list[idx] == 'magenta':
                symbol.setColor(QColor(255, 0, 255, 255))
            elif color_list[idx] == 'mediumturquoise':
                symbol.setColor(QColor(72, 209, 204, 255))
            elif color_list[idx] == 'yellow':
                symbol.setColor(QColor(255, 215, 0, 255))
            elif color_list[idx] == 'lime':
                symbol.setColor(QColor(0, 255, 0, 255))
            elif color_list[idx] == 'black':
                symbol.setColor(QColor(0, 0, 0, 255))
            elif color_list[idx] == 'grey':
                symbol.setColor(QColor(120, 120, 120, 255))
            QgsProject.instance().addMapLayer(layer)

        ## Add memory layers ##
        # Polygon Layer
        self.polygonLayer = QgsVectorLayer(
            "Polygon?crs=EPSG:27700", "polygonLayer", "memory")
        s = QgsLineSymbol.createSimple({"color": "blue", "width": "1"})
        r = QgsSingleSymbolRenderer(s)
        self.polygonLayer.setRenderer(r)
        symbol = self.polygonLayer.renderer().symbol()
        symbol.setColor(QColor(Qt.red))
        QgsProject.instance().addMapLayer(self.polygonLayer)

        # Line Layer
        self.lineLayer = QgsVectorLayer(
            "LineString?crs=EPSG:27700", "lineLayer", "memory")
        s = QgsLineSymbol.createSimple({"color": "blue", "width": "1"})
        r = QgsSingleSymbolRenderer(s)
        self.lineLayer.setRenderer(r)
        symbol = self.lineLayer.renderer().symbol()
        symbol.setColor(QColor(Qt.red))
        QgsProject.instance().addMapLayer(self.lineLayer)

        # Point Layer
        self.pointLayer = QgsVectorLayer(
            "Point?crs=EPSG:27700", "pointLayer", "memory")
        s = QgsLineSymbol.createSimple({"color": "blue", "width": "1"})
        r = QgsSingleSymbolRenderer(s)
        self.pointLayer.setRenderer(r)
        symbol = self.pointLayer.renderer().symbol()
        symbol.setColor(QColor(Qt.red))
        QgsProject.instance().addMapLayer(self.pointLayer)

        self.createOsLayers()
        self.addSatelliteBaseMap()
        self.showVisibleMapLayers()
        self.mapCanvas.setCenter(self.center)
        self.mapCanvas.zoomScale(200)
        self.mapCanvas.refresh()

    def showBasemap1Layer(self):
        self.showVisibleMapLayers()

    def showBasemap2Layer(self):
        self.showVisibleMapLayers()

    def showBasemap3Layer(self):
        self.showVisibleMapLayers()

    def showVisibleMapLayers(self):
        """Setup layers visibility according to map scale and other application options"""

        scaleDenominator = self.mapCanvas.scale()

        if scaleDenominator <= 1.0:
            return

        layers = []

        if scaleDenominator <= MAX_SCALE_DENOMINATOR_FOR_BASEMAP:
            # Use OS layer or Satellite depending on the status of the Satellite toggle button
            if self.actionSatelliteBaseMap.isChecked():
                layers.append(self.satelliteBaseMap)
            else:
                layers.append(self.osBaseLayer)
        else:
            layers.append(QgsProject.instance(
            ).mapLayersByName('OpenStreetMap')[0])

        if scaleDenominator <= MAX_SCALE_DENOMINATOR_FOR_FEATURES:
            layers.append(self.osBuildingsLayer)

        house_layer_name = self.housename + '_%s' % self.layer_num
        for layer_name in (house_layer_name, "polygonLayer",
                           "lineLayer", "pointLayer"):
            try:
                layers.append(QgsProject.instance().mapLayersByName(layer_name)[0])
            except BaseException:
                logger.debug("Layer not found: %s" % layer_name)

        layers.reverse()

        if set(layers) != set(self.mapCanvas.layers()):
            self.mapCanvas.setLayers(layers)
            self.mapCanvas.refresh()
            logger.debug("Refreshing canvas...")

    def setFalse(self):
        self.actionOH.setChecked(False)
        self.actionG.setChecked(False)
        self.actionF.setChecked(False)
        self.actionE1.setChecked(False)
        self.actionE2.setChecked(False)
        self.actionA.setChecked(False)
        self.actionL.setChecked(False)
        self.actionAP.setChecked(False)

    def setOriginalHouse(self):
        if self.actionOH.isChecked():
            self.setFalse()
            self.actionOH.setChecked(True)
            self.layer_num = 1
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setGroundMap(self):
        if self.actionG.isChecked():
            self.setFalse()
            self.actionG.setChecked(True)
            self.layer_num = 2
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setFirstMap(self):
        if self.actionF.isChecked():
            self.setFalse()
            self.actionF.setChecked(True)
            self.layer_num = 3
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setExt1Map(self):
        if self.actionE1.isChecked():
            self.setFalse()
            self.actionE1.setChecked(True)
            self.layer_num = 4
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setExt2Map(self):
        if self.actionE2.isChecked():
            self.setFalse()
            self.actionE2.setChecked(True)
            self.layer_num = 5
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setAtticMap(self):
        if self.actionA.isChecked():
            self.setFalse()
            self.actionA.setChecked(True)
            self.layer_num = 6
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setLineMap(self):
        if self.actionL.isChecked():
            self.setFalse()
            self.actionL.setChecked(True)
            self.layer_num = 7
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def setAllPlanningMap(self):
        if self.actionAP.isChecked():
            self.setFalse()
            self.actionAP.setChecked(True)
            self.layer_num = 8
            self.setupMapTools()
            self.showVisibleMapLayers()
        else:
            self.setFalse()
            self.layer_num = 9
            self.setupMapTools()
            self.showVisibleMapLayers()

    def closeEvent(self, event):
        self.quit()

    def checkSaveProject(self):
        """Makes sure the pending commits are saved before quitting or changing project

        Returns True unless Cancel was chosen in the confirmation dialog.

        """
        if self.editing and self.modified:
            reply = QMessageBox.question(self, "Confirm", "Save Changes?",
                                         QMessageBox.Yes |
                                         QMessageBox.No |
                                         QMessageBox.Cancel,
                                         QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                for l in self.houseLayers():
                    l.commitChanges()
            elif reply == QMessageBox.No:
                for l in self.houseLayers():
                    l.rollBack()

            return reply != QMessageBox.Cancel

        else:
            return True

    def quit(self):
        """Quit from the main dialog"""

        if self.checkSaveProject():
            self.close()

    def zoomIn(self):
        self.mapCanvas.zoomIn()

    def zoomOut(self):
        self.mapCanvas.zoomOut()

    def setPanMode(self):
        self.mapCanvas.setMapTool(self.panTool)

    def get_config(self, key, default=''):
        # return a config parameter
        return self.config.value(key, default)

    def set_config(self, key, value):
        # set a config parameter
        return self.config.setValue(key, value)

    def process_point(self, place, point):
        # lon lat and transform
        point = QgsPoint(point[1], point[0])
        crs = QgsCoordinateReferenceSystem(4326)
        point = pointFromWGS84(point, crs)

        # Set the extent to our new rectangle
        self.mapCanvas.setCenter(point)

        scale = 100.0
        # adjust scale to display correct scale in qgis
        self.mapCanvas.zoomScale(scale)

        # Refresh the map
        self.mapCanvas.refresh()
        # save point
        self.save_point(point, unicode(place))

    def setupMapToolsHouseLayer(self, houseLayer):

        self.addPointTool = AddPointTool(
            self.mapCanvas, houseLayer, self.geometryAdded, self.onHouseAdded, self.layer_num)

        self.addLineTool = CaptureTool(
            self.mapCanvas, houseLayer, self.geometryAdded, CaptureTool.CAPTURE_LINE, self.onHouseAdded, self.layer_num)

        def _onVertexMoved(rubberBand):
            """Internal signal handler"""

            vertex_idx = rubberBand.partSize(rubberBand.size() - 1) - 1
            last_point = rubberBand.getPoint(rubberBand.size() - 1, vertex_idx)
            previous_point = rubberBand.getPoint(
                rubberBand.size() - 1, vertex_idx - 1)
            self.updateStatusBar(last_point.distance(previous_point))

        self.addLineTool.vertexMoved.connect(_onVertexMoved)

        self.addPolygonTool = CaptureTool(
            self.mapCanvas, houseLayer, self.geometryAdded, CaptureTool.CAPTURE_POLYGON, self.onHouseAdded,
            self.layer_num)

        self.addPolygonTool.vertexMoved.connect(_onVertexMoved)

        self.deletePointTool = DeleteTool(
            self.mapCanvas, houseLayer, self.geometryDeleted)
        self.deleteLineTool = DeleteTool(
            self.mapCanvas, houseLayer, self.geometryDeleted)
        self.deletePolygonTool = DeleteTool(
            self.mapCanvas, houseLayer, self.geometryDeleted)
        self.editPointTool = MovePointTool(
            self.mapCanvas, houseLayer, self.onGeometryChanged)
        self.editLineTool = EditTool(
            self.mapCanvas, houseLayer, self.onGeometryChanged)
        self.editPolygonTool = EditTool(
            self.mapCanvas, houseLayer, self.onGeometryChanged)
        self.getInfoTool = GetInfoTool(
            self.mapCanvas, houseLayer, self.onGetInfo, self.geometryDeleted, self.geometryAdded,
            CaptureTool.CAPTURE_POLYGON, self.onHouseAdded, self.layer_num)
        self.splitTool = SplitTool(
            self.mapCanvas, houseLayer, self.onSplitHouse, self.geometryDeleted, self.geometryAdded,
            CaptureTool.CAPTURE_POLYGON, self.onHouseAdded)
        self.manualTileTool = ManualTile(
            self.mapCanvas, houseLayer, self.onManualTile, self.geometryDeleted, self.geometryAdded,
            CaptureTool.CAPTURE_POLYGON, self.onHouseAdded, self.type_list, self.layer_num)
        self.manualLoftTileTool = ManualLoftTile(
            self.mapCanvas, houseLayer, self.onManualLoftTile, self.geometryDeleted, self.geometryAdded,
            CaptureTool.CAPTURE_POLYGON, self.onHouseAdded, self.type_list, self.layer_num)

    def setupMapTools(self):

        self.panTool = PanTool(self.mapCanvas)
        self.panTool.setAction(self.actionPan)

        house_layer = getattr(self, 'houseLayer%s' % self.layer_num)
        self.setupMapToolsHouseLayer(house_layer)
        self.addPointTool.setAction(self.actionAdd)
        self.addLineTool.setAction(self.actionAdd)
        self.addPolygonTool.setAction(self.actionAdd)
        self.deletePointTool.setAction(self.actionDelete)
        self.deleteLineTool.setAction(self.actionDelete)
        self.deletePolygonTool.setAction(self.actionDelete)
        self.editPointTool.setAction(self.actionEdit)
        self.editLineTool.setAction(self.actionEdit)
        self.editPolygonTool.setAction(self.actionEdit)
        self.getInfoTool.setAction(self.actionGetInfo)
        self.splitTool.setAction(self.actionSplitPoly)
        self.manualTileTool.setAction(self.actionManualTileExt)
        self.manualLoftTileTool.setAction(self.actionManualTileLoft)

    def adjustActions(self):
        """ Enable or disable actions, as appropriate.
        """

        if self.editing:
            logger.debug('editing active')
            self.actionAdd.setEnabled(True)
            self.actionEdit.setEnabled(True)
            self.actionDelete.setEnabled(True)
            self.actionGetInfo.setEnabled(True)
            # self.actionPDF.setEnabled(False)
            self.actionSelectFeature.setEnabled(True)
            self.actionImportFeatures.setEnabled(True)
            self.actionSplitPoly.setEnabled(True)
            self.actionManualTileExt.setEnabled(True)
            self.actionManualTileLoft.setEnabled(True)
        else:
            logger.debug('editing not active')
            self.actionAdd.setEnabled(False)
            self.actionEdit.setEnabled(False)
            self.actionDelete.setEnabled(False)
            self.actionGetInfo.setEnabled(False)
            # self.actionPDF.setEnabled(True)
            self.actionSelectFeature.setEnabled(False)
            self.actionImportFeatures.setEnabled(False)
            self.actionSplitPoly.setEnabled(False)
            self.actionManualTileExt.setEnabled(False)
            self.actionManualTileLoft.setEnabled(False)

        curTool = self.mapCanvas.mapTool()

        self.actionPan.setChecked(curTool == self.panTool)
        self.actionAdd.setChecked(curTool == self.addLineTool)
        self.actionAdd.setChecked(curTool == self.addPolygonTool)
        self.actionEdit.setChecked(curTool == self.editPolygonTool)
        self.actionEdit.setChecked(curTool == self.editLineTool)
        self.actionDelete.setChecked(curTool == self.deletePolygonTool)
        self.actionDelete.setChecked(curTool == self.deleteLineTool)
        self.actionGetInfo.setChecked(curTool == self.getInfoTool)
        self.actionSplitPoly.setChecked(curTool == self.splitTool)
        self.actionManualTileExt.setChecked(curTool == self.manualTileTool)
        self.actionManualTileLoft.setChecked(curTool == self.manualLoftTileTool)

    def getInfo(self):
        if not self.actionGetInfo.isChecked():
            self.setPanMode()
            return
        else:
            self.mapCanvas.setMapTool(self.getInfoTool)

    def graphFunction(self):
        if self.actionGraph.isChecked():
            # define a global house geometry object
            gt = Geometry()

            # Import data from databases
            r = Reader(self.dropbox_foldername, gt, self.datum, app_bool=True)
            plot_bool, plot_color_bool, move_bool, points_bool, plot_PD, plot_PP, plot_AP = True, False, True, False, True, False, False
            main_pd, main_pp, main_ap, left, right = r.run_reader(r, plot_bool, plot_color_bool, move_bool, points_bool,
                                                                  plot_PD, plot_PP, plot_AP)
            self.actionGraph.setChecked(False)
        else:
            self.actionGraph.setChecked(False)

    def graphFunctionPP(self):
        if self.actionGraphPP.isChecked():
            # define a global house geometry object
            gt = Geometry()
            # Import data from databases
            r = Reader(self.dropbox_foldername, gt, self.datum, app_bool=True)
            plot_bool, plot_color_bool, move_bool, points_bool, plot_PD, plot_PP, plot_AP = True, False, True, False, False, True, False
            main_pd, main_pp, main_ap, left, right = r.run_reader(r, plot_bool, plot_color_bool, move_bool, points_bool,
                                                                  plot_PD, plot_PP, plot_AP)
            self.actionGraphPP.setChecked(False)
        else:
            self.actionGraphPP.setChecked(False)

    def graphFunctionAP(self):
        if self.actionGraphAP.isChecked():
            # define a global house geometry object
            gt = Geometry()

            # Import data from databases
            r = Reader(self.dropbox_foldername, gt, self.datum, app_bool=True)
            plot_bool, plot_color_bool, move_bool, points_bool, plot_PD, plot_PP, plot_AP = True, False, True, False, False, False, True
            main_pd, main_pp, main_ap, left, right = r.run_reader(r, plot_bool, plot_color_bool, move_bool, points_bool,
                                                                  plot_PD, plot_PP, plot_AP)
            self.actionGraphAP.setChecked(False)
        else:
            self.actionGraphAP.setChecked(False)

    def displayScrape(self):
        # self.saveCurrLayer()
        if self.scrapeNotDone:
            try:
                print(self.datum)
            except:
                self.datum = {
                    'id': '',
                    'address': '',
                    'link': '',
                    'description': '',
                    'date': '',
                    'price': '',
                    'type house': '',
                    'conservation_area': '',
                    'green_belt': '',
                    'image_url': '',
                    'bedrooms': '',
                    'bathrooms': '',
                    'reception': '',
                    'area': '',
                    'Borough': '',
                    'PPSQM': '',
                    'BPSQM': '',
                    'summary_report_bool': False,
                    'add_value_bool': False,
                    'n_rooms': 0
                }
        check, accepted = self.checkwebscrape(self.datum)
        self.datum = check

        if accepted:
            self.actionShowScrape.setChecked(False)
        else:
            self.actionShowScrape.setChecked(False)

    def plannerNotes(self):
        dialog = ManualPlannerNotesDialog(parent=self)
        dialog.input_foldername(self.dropbox_foldername)
        dialog.restoreSettings()
        accepted = False
        if dialog.exec() == QDialog.Accepted:
            dialog.saveSettings()
            accepted = True

        if accepted:
            self.actionPlannerNotes.setChecked(False)
        else:
            self.actionPlannerNotes.setChecked(False)

    def saveCurrLayer(self):

        future = self.now + 100
        if time.time() >= future:
            for l in self.houseLayers():
                l.commitChanges()
                self.now = time.time()

    def manualRules(self):
        gt = Geometry()
        r = Reader(self.dropbox_foldername, gt, self.datum, app_bool=True)
        dialog = ManualRulesDialog(self, r, gt)
        if dialog.exec() == QDialog.Accepted:
            self.actionManualRules.setChecked(False)
        else:
            self.actionManualRules.setChecked(False)

    def manualMeasures(self):
        gt = Geometry()
        r = Reader(self.dropbox_foldername, gt, self.datum, app_bool=True)
        dialog = ManualMeasuresInfoDialog(self, r, gt)
        dialog.input_foldername(self.dropbox_foldername)
        dialog.restoreSettings()
        if dialog.exec() == QDialog.Accepted:
            dialog.saveSettings()
            try:
                HW = float(dialog.HW.text())
            except BaseException:
                HW = -100.0
            try:
                HD = float(dialog.HD.text())
            except BaseException:
                HD = -100.0
            try:
                SW = float(dialog.SW.text())
            except BaseException:
                SW = -100.0
            try:
                SD = float(dialog.SD.text())
            except BaseException:
                SD = -100.0
            try:
                SDR = float(dialog.SDR.text())
            except BaseException:
                SDR = -100.0
            try:
                RDGL = float(dialog.RDGL.text())
            except BaseException:
                RDGL = -100.0
            try:
                GB1 = float(dialog.GB1.text())
            except BaseException:
                GB1 = -100.0
            try:
                GFB1 = float(dialog.GFB1.text())
            except BaseException:
                GFB1 = -100.0
            try:
                GSE1 = float(dialog.GSE1.text())
            except BaseException:
                GSE1 = -100.0
            try:
                GFN1 = float(dialog.GFN1.text())
            except BaseException:
                GFN1 = -100.0
            try:
                FFN1 = float(dialog.FFN1.text())
            except BaseException:
                FFN1 = -100.0
            try:
                GFNF1 = float(dialog.GFNF1.text())
            except BaseException:
                GFNF1 = -100.0
            try:
                FFNF1 = float(dialog.FFNF1.text())
            except BaseException:
                FFNF1 = -100.0
            try:
                GB2 = float(dialog.GB2.text())
            except BaseException:
                GB2 = -100.0
            try:
                GFB2 = float(dialog.GFB2.text())
            except BaseException:
                GFB2 = -100.0
            try:
                GSE2 = float(dialog.GSE2.text())
            except BaseException:
                GSE2 = -100.0
            try:
                GFN2 = float(dialog.GFN2.text())
            except BaseException:
                GFN2 = -100.0
            try:
                FFN2 = float(dialog.FFN2.text())
            except BaseException:
                FFN2 = -100.0
            try:
                GFNF2 = float(dialog.GFNF2.text())
            except BaseException:
                GFNF2 = -100.0
            try:
                FFNF2 = float(dialog.FFNF2.text())
            except BaseException:
                FFNF2 = -100.0
            try:
                H = float(dialog.H.text())
            except BaseException:
                H = 3.0
            try:
                H1 = float(dialog.H1.text())
            except BaseException:
                H1 = 3.0
            try:
                HA = float(dialog.HA.text())
            except BaseException:
                HA = 3.0
            try:
                HO3 = float(dialog.HO3.text())
            except BaseException:
                HO3 = 3.0
            try:
                HO4 = float(dialog.HO4.text())
            except BaseException:
                HO4 = 3.0
            try:
                HO5 = float(dialog.HO5.text())
            except BaseException:
                HO5 = 3.0
            try:
                HO6 = float(dialog.HO6.text())
            except BaseException:
                HO6 = 3.0
            try:
                HO19 = float(dialog.HO19.text())
            except BaseException:
                HO19 = 3.0
            try:
                HO20 = float(dialog.HO20.text())
            except BaseException:
                HO20 = 3.0
            try:
                RHO3 = float(dialog.RHO3.text())
            except BaseException:
                RHO3 = 1.0
            try:
                RHO4 = float(dialog.RHO4.text())
            except BaseException:
                RHO4 = 1.0
            try:
                RHO5 = float(dialog.RHO5.text())
            except BaseException:
                RHO5 = 1.0
            try:
                RHO6 = float(dialog.RHO6.text())
            except BaseException:
                RHO6 = 1.0
            try:
                RHO19 = float(dialog.RHO19.text())
            except BaseException:
                RHO19 = 1.0
            try:
                RHO20 = float(dialog.RHO20.text())
            except BaseException:
                RHO20 = 1.0

            str1 = self.dropbox_foldername + "/measures.pickle"
            b = {'HW': HW, 'HD': HD, 'SW': SW, 'SD': SD, 'SDR': SDR, 'RDGL': RDGL,
                 'GB1': GB1, 'GFB1': GFB1, 'GSE1': GSE1, 'GFN1': GFN1, 'FFN1': FFN1, 'GFNF1': GFNF1, 'FFNF1': FFNF1,
                 'GB2': GB2, 'GFB2': GFB2, 'GSE2': GSE2, 'GFN2': GFN2, 'FFN2': FFN2, 'GFNF2': GFNF2, 'FFNF2': FFNF2,
                 'H': H, 'H1': H1, 'HL': HA, 'HO3': HO3, 'HO4': HO4, 'HO5': HO5, 'HO6': HO6, 'HO19': HO19, 'HO20': HO20,
                 'RHO3': RHO3, 'RHO4': RHO4, 'RHO5': RHO5, 'RHO6': RHO6, 'RHO19': RHO19, 'RHO20': RHO20}

            with open(str1, 'wb') as handle:
                pickle.dump(b, handle, protocol=pickle.HIGHEST_PROTOCOL)
            self.actionManualMeasures.setChecked(False)
        else:
            self.actionManualMeasures.setChecked(False)

    def manualExtensions(self):

        dialog = ManualExtensionsInfoDialog(self)
        dialog.input_foldername(self.dropbox_foldername)
        dialog.restoreSettings()

        if dialog.exec() == QDialog.Accepted:
            dialog.saveSettings()
            for i in range(len(self.Ext_list)):
                if dialog.E[i].isChecked():
                    self.E[i] = True
                else:
                    self.E[i] = False
                self.Ext_list[i] = dialog.Ext_list[i]
                if dialog.E_PP[i].isChecked():
                    self.E_PP[i] = True
                else:
                    self.E_PP[i] = False
                self.Ext_PP_list[i] = dialog.Ext_PP_list[i]
                if dialog.E_AP[i].isChecked():
                    self.E_AP[i] = True
                else:
                    self.E_AP[i] = False
                self.Ext_AP_list[i] = dialog.Ext_AP_list[i]

            str1 = self.dropbox_foldername + "/extensions.pickle"
            a = {'E': self.E, 'E_PP': self.E_PP, 'E_AP': self.E_AP, 'Ext_list': self.Ext_list,
                 'Ext_PP_list': self.Ext_PP_list,
                 'Ext_AP_list': self.Ext_AP_list, 'Push_to_combo': self.E[-1]}

            with open(str1, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

            if dialog.E[-1].isChecked():
                self.fillCombosExt()

            self.actionManualExtensions.setChecked(False)
        else:
            self.actionManualExtensions.setChecked(False)

    def manualLoft(self):

        dialog = ManualLoftInfoDialog(self)
        dialog.input_foldername(self.dropbox_foldername)
        dialog.restoreSettings()

        if dialog.exec() == QDialog.Accepted:
            dialog.saveSettings()
            for i in range(len(self.Ext_loft_list)):
                if dialog.EL[i].isChecked():
                    self.EL[i] = True
                else:
                    self.EL[i] = False
                if dialog.EL_PP[i].isChecked():
                    self.EL_PP[i] = True
                else:
                    self.EL_PP[i] = False
                if dialog.EL_AP[i].isChecked():
                    self.EL_AP[i] = True
                else:
                    self.EL_AP[i] = False
                if i > 0:
                    self.Ext_loft_list[i] = dialog.Ext_loft_list[i]

            str1 = self.dropbox_foldername + "/loft.pickle"
            a = {'EL': self.EL, 'EL_PP': self.EL_PP, 'EL_AP': self.EL_AP, 'Ext_loft_list': self.Ext_loft_list,
                 'Push_to_combo': self.EL[-1]}

            with open(str1, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

            if dialog.EL[-1].isChecked():
                self.fillCombosLoft()

            self.actionManualLoft.setChecked(False)
        else:
            self.actionManualLoft.setChecked(False)

    def logDialog(self):

        log = loggerHelpDialog(self)

        if log.exec() == QDialog.Accepted:
            pass
        else:
            pass

    def manualCombo(self):

        dialog = ManualComboInfoDialog(self)
        dialog.input_foldername(self.dropbox_foldername)
        dialog.restoreSettings()

        if dialog.exec() == QDialog.Accepted:
            dialog.saveSettings()

            self.gf_manual = dialog.GF.isChecked()
            self.ff_manual = dialog.FF.isChecked()
            self.at_manual = dialog.AT.isChecked()
            self.gf_manual_combo_list = dialog.GF_list.text()
            self.ff_manual_combo_list = dialog.FF_list.text()
            self.at_manual_combo_list = dialog.AT_list.text()

            self.gf_PP_manual = dialog.GF_PP.isChecked()
            self.ff_PP_manual = dialog.FF_PP.isChecked()
            self.at_PP_manual = dialog.AT_PP.isChecked()
            self.gf_PP_manual_combo_list = dialog.GF_list_PP.text()
            self.ff_PP_manual_combo_list = dialog.FF_list_PP.text()
            self.at_PP_manual_combo_list = dialog.AT_list_PP.text()

            self.gf_AP_manual = dialog.GF_AP.isChecked()
            self.ff_AP_manual = dialog.FF_AP.isChecked()
            self.at_AP_manual = dialog.AT_AP.isChecked()
            self.gf_AP_manual_combo_list = dialog.GF_list_AP.text()
            self.ff_AP_manual_combo_list = dialog.FF_list_AP.text()
            self.at_AP_manual_combo_list = dialog.AT_list_AP.text()

            str1 = self.dropbox_foldername + "/combos.pickle"
            a = {'GF': self.gf_manual, 'GF_list': self.gf_manual_combo_list, 'FF': self.ff_manual,
                 'FF_list': self.ff_manual_combo_list, 'AT': self.at_manual, 'AT_list': self.at_manual_combo_list,
                 'GF_PP': self.gf_PP_manual, 'GF_list_PP': self.gf_PP_manual_combo_list, 'FF_PP': self.ff_PP_manual,
                 'FF_list_PP': self.ff_PP_manual_combo_list, 'AT_PP': self.at_PP_manual,
                 'AT_list_PP': self.at_PP_manual_combo_list,
                 'GF_AP': self.gf_AP_manual, 'GF_list_AP': self.gf_AP_manual_combo_list, 'FF_AP': self.ff_AP_manual,
                 'FF_list_AP': self.ff_AP_manual_combo_list, 'AT_AP': self.at_AP_manual,
                 'AT_list_AP': self.at_AP_manual_combo_list}

            with open(str1, 'wb') as handle:
                pickle.dump(a, handle, protocol=pickle.HIGHEST_PROTOCOL)

            if dialog.Push_to_Extensions_Loft.isChecked():
                print('Push to Extensions and Loft')

            self.actionManualCombo.setChecked(False)
        else:
            self.actionManualCombo.setChecked(False)

    def modTrimCombosExt(self, list, list_names, pd_bool):

        keys = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '18', '19', '20', '21', '22', '23',
                '3a', '4a', '5a', '6a', '19a', '20a']
        count = 0
        possible_extensions = {}
        for k in keys:
            # count = count + 1
            possible_extensions[k] = False

        for k in keys:
            count = count + 1
            # possible_extensions[k] = False
            if list[count] == True:
                k_ = list_names[count][4:]
                possible_extensions[k_] = True

        ext = Extensions(None, None, pd_bool, True)
        combo_gf, combo_ff, combo_at, extensions = ext.combo_gf, ext.combo_ff, ext.combo_at, ext.extensions
        combo_gf, combo_ff = ext.trim_combo_by_possible_ext(combo_gf, combo_ff, possible_extensions)
        if len(combo_gf) > 0:
            combo_gf.sort(key=len)
            combo_gf.reverse()
        if len(combo_ff) > 0:
            combo_ff.sort(key=len)
            combo_ff.reverse()

        return combo_gf, combo_ff

    def modTrimCombosLoft(self, list, list_names, pd_bool):

        gt = Geometry()
        r = Reader(self.dropbox_foldername, gt, app_bool=True)
        plot_bool, plot_color_bool, move_bool, points_bool, plot_PD, plot_PP, plot_AP = False, False, True, False, False, False, False
        main_pd, main_pp, main_ap, left, right = r.run_reader(r, plot_bool, plot_color_bool, move_bool, points_bool,
                                                              plot_PD, plot_PP, plot_AP)

        keys = ['1L', '2L', '3L', '4L', '5L', '6L', '7L', '8L']
        count = 0
        possible_extensions = {}
        for k in keys:
            count = count + 1
            possible_extensions[k] = False
            if list[count] == True:
                possible_extensions[k] = True
        ext = Extensions(r, main_pd, pd_bool, True)
        combo_gf, combo_ff, combo_at, extensions = ext.combo_gf, ext.combo_ff, ext.combo_at, ext.extensions
        combo_at = ext.trim_combo_by_possible_loft(combo_at, possible_extensions)

        if len(combo_at) > 0:
            combo_at.sort(key=len)
            combo_at.reverse()
        return combo_at

    def fillCombosExt(self):

        self.pickle_filename = self.dropbox_foldername + "/extensions.pickle"
        extensionList_pd = []
        extensionList_pp = []
        extensionList_ap = []
        extensionList_pd_names = []
        extensionList_pp_names = []
        extensionList_ap_names = []

        file = open(self.pickle_filename, 'rb')
        data = pickle.load(file)
        file.close()
        self.pushBoolean = data['Push_to_combo']
        if data['Push_to_combo'] == True:
            for i in range(len(self.Ext_list)):
                extensionList_pd.append(data['E'][i])
                extensionList_pp.append(data['E_PP'][i])
                extensionList_ap.append(data['E_AP'][i])
                extensionList_pd_names.append(data['Ext_list'][i])
                extensionList_pp_names.append(data['Ext_PP_list'][i])
                extensionList_ap_names.append(data['Ext_AP_list'][i])

        # Fill in PD data to combos.pickle
        combo_gf, combo_ff = self.modTrimCombosExt(extensionList_pd, extensionList_pd_names, True)
        combo_gf_pp, combo_ff_pp = self.modTrimCombosExt(extensionList_pp, extensionList_pp_names, False)
        combo_gf_ap, combo_ff_ap = self.modTrimCombosExt(extensionList_ap, extensionList_ap_names, False)

        self.pickle_filename = self.dropbox_foldername + "/combos.pickle"
        file = open(self.pickle_filename, 'rb')
        data = pickle.load(file)
        if len(combo_gf) > 0:
            data['GF_list'] = self.list_of_list_to_str(combo_gf)
            data['FF_list'] = self.list_of_list_to_str(combo_ff)
        if len(data['GF_list']) > 0:
            data['GF'] = True
        if len(data['FF_list']) > 0:
            data['FF'] = True

        if len(combo_gf_pp) > 0:
            data['GF_list_PP'] = self.list_of_list_to_str(combo_gf_pp)
            data['FF_list_PP'] = self.list_of_list_to_str(combo_ff_pp)
        if len(data['GF_list_PP']) > 0:
            data['GF_PP'] = True
        if len(data['FF_list_PP']) > 0:
            data['FF_PP'] = True

        if len(combo_gf_ap) > 0:
            data['GF_list_AP'] = self.list_of_list_to_str(combo_gf_ap)
            data['FF_list_AP'] = self.list_of_list_to_str(combo_ff_ap)
        if len(data['GF_list_AP']) > 0:
            data['GF_AP'] = True
        if len(data['FF_list_AP']) > 0:
            data['FF_AP'] = True

        with open(self.pickle_filename, 'wb') as handle:
            pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)

    def fillCombosLoft(self):

        self.pickle_filename = self.dropbox_foldername + "/loft.pickle"
        loftList_pd = []
        loftList_pp = []
        loftList_ap = []
        loftList_pd_names = []
        loftList_pp_names = []
        loftList_ap_names = []

        file = open(self.pickle_filename, 'rb')
        data = pickle.load(file)
        file.close()
        self.pushBooleanLoft = data['Push_to_combo']
        if data['Push_to_combo'] == True:
            for i in range(len(self.Ext_loft_list)):
                loftList_pd.append(data['EL'][i])
                loftList_pp.append(data['EL_PP'][i])
                loftList_ap.append(data['EL_AP'][i])
                loftList_pd_names.append(data['Ext_loft_list'][i])
                loftList_pp_names.append(data['Ext_loft_list'][i])
                loftList_ap_names.append(data['Ext_loft_list'][i])
        combo_at = self.modTrimCombosLoft(loftList_pd, loftList_pd_names, True)
        combo_at_pp = self.modTrimCombosLoft(loftList_pp, loftList_pp_names, False)
        combo_at_ap = self.modTrimCombosLoft(loftList_ap, loftList_pp_names, False)

        self.pickle_filename = self.dropbox_foldername + "/combos.pickle"
        file = open(self.pickle_filename, 'rb')
        data = pickle.load(file)
        file.close()
        if len(combo_at) > 0:
            data['AT_list'] = self.list_of_list_to_str(combo_at)
        if len(data['AT_list']) > 0:
            data['AT'] = True

        if len(combo_at_pp) > 0:
            data['AT_list_PP'] = self.list_of_list_to_str(combo_at_pp)
        if len(data['AT_list_PP']) > 0:
            data['AT_PP'] = True

        if len(combo_at_ap) > 0:
            data['AT_list_AP'] = self.list_of_list_to_str(combo_at_ap)
        if len(data['AT_list_AP']) > 0:
            data['AT_AP'] = True

        with open(self.pickle_filename, 'wb') as handle:
            pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)

    def list_of_list_to_str(self, list):
        list_str = ''
        list_str += '['
        for l in list:
            list_str += '['
            for s in range(len(l)):
                list_str += "'" + l[s] + "'" + ','
            if len(l) > 0:
                list_str = list_str[:-1]
            list_str += '],'
        list_str = list_str[:-1]
        list_str += ']'
        return list_str

    def manualComboPP(self):
        pass

    def returnCouncil(self):
        try:
            time.sleep(3)
            pds = property_data_scraper(self.address, 0, '', '', '', 4, 4, '', '')
            return pds.postcode_to_council()['council']
        except:
            return 'None'

    def report(self):
        gt = Geometry()
        council = self.returnCouncil()
        r = Reader(self.dropbox_foldername, gt, app_bool=True)
        plot_bool, plot_color_bool, move_bool, points_bool, plot_PD, plot_PP, plot_AP = False, False, True, False, False, False, False
        main_pd, main_pp, main_ap, left, right = r.run_reader(r, plot_bool, plot_color_bool, move_bool, points_bool,
                                                              plot_PD, plot_PP, plot_AP)

        dialog = ReportInfoDialog(self, reader=r, layer_num=self.layer_num, council=council,
                                  dropbox_foldername=self.dropbox_foldername, datum=self.datum)
        dialog.exec()
        self.actionReport.setChecked(False)

    def webscrape(self):
        webscrape = ManualURLInfoDialog(parent=self, dropbox_foldername=self.dropbox_foldername, address=self.address)
        if webscrape.exec() == QDialog.Accepted:
            webscrape.saveSettings()
            url = webscrape.url.text()
            process = ProcessSingleHouse(self.address, url, self.id_list)
            if webscrape.check:
                process.get_data_from_url()
                process.get_property_data_from_address()
            return process
        else:
            return None

    def checkwebscrape(self, _datum):
        checkwebscraper = ManualCheckURLInfoDialog(parent=self, datum=_datum, id_list=self.id_list)
        checkwebscraper.input_foldername(self.dropbox_foldername)
        if checkwebscraper.exec() == QDialog.Accepted:
            return checkwebscraper.datum, True
        else:
            return None, False

    def requestedExtensions(self):
        pass

    def makePDF(self):
        pass
        # result = rs.makepdf()

    def houseLayers(self):
        """Returns a list of all house layers"""

        return [getattr(self, 'houseLayer%s' % idx)
                for idx in range(1, NUMBER_OF_SQLITE_HOUSE_LAYERS + 1)]

    def _setLayerEditMode(self, mode):
        """Mode can be "line" or "polygon"""

        if mode == "line":
            self.actionEditLine.setChecked(True)
            self.actionEditPolygon.setChecked(False)
            self.curEditedLayer = self.lineLayer
        else:
            self.actionEditLine.setChecked(False)
            self.actionEditPolygon.setChecked(True)
            self.curEditedLayer = self.polygonLayer

        # Common code: manage all house layers
        if self.editing:

            # Stop editing the house layers.
            if self.modified:
                reply = QMessageBox.question(self, "Confirm", "Save Changes?",
                                             QMessageBox.Yes | QMessageBox.No,
                                             QMessageBox.Yes)
                if reply == QMessageBox.Yes:
                    for l in self.houseLayers():
                        l.commitChanges()

                else:
                    for l in self.houseLayers():
                        l.rollBack()

            else:
                for l in self.houseLayers():
                    l.commitChanges()

            for l in self.houseLayers():
                l.triggerRepaint()

            if mode == "line":
                self.actionEditLine.setChecked(False)
            else:
                self.actionEditPolygon.setChecked(False)
            self.curEditedLayer = None
            self.setPanMode()
            self.editing = False

        else:
            # Start editing the house layers.
            for l in self.houseLayers():
                l.startEditing()

            self.modified = False
            self.editing = True

        self.adjustActions()

    def setEditPolygonMode(self):
        """Set polygon editing mode"""

        self._setLayerEditMode("polygon")

    def setEditLineMode(self):
        """Set linestring editing mode"""

        self._setLayerEditMode("line")

    def addFeature(self):
        if not self.actionAdd.isChecked():
            self.setPanMode()
            return

        if self.curEditedLayer == self.pointLayer:
            tool = self.addPointTool
        elif self.curEditedLayer == self.lineLayer:
            tool = self.addLineTool
        elif self.curEditedLayer == self.polygonLayer:
            tool = self.addPolygonTool
        self.mapCanvas.setMapTool(tool)

    def editFeature(self):
        if not self.actionEdit.isChecked():
            self.setPanMode()
            return

        if self.curEditedLayer == self.pointLayer:
            tool = self.editPointTool
        elif self.curEditedLayer == self.lineLayer:
            tool = self.editLineTool
        elif self.curEditedLayer == self.polygonLayer:
            tool = self.editPolygonTool
        self.mapCanvas.setMapTool(tool)

    def deleteFeature(self):
        if not self.actionDelete.isChecked():
            self.setPanMode()
            return

        if self.curEditedLayer == self.pointLayer:
            tool = self.deletePointTool
        elif self.curEditedLayer == self.lineLayer:
            tool = self.deleteLineTool
        elif self.curEditedLayer == self.polygonLayer:
            tool = self.deletePolygonTool
        self.mapCanvas.setMapTool(tool)

    def splitPoly(self):
        if not self.actionSplitPoly.isChecked():
            self.setPanMode()
            return
        else:
            self.mapCanvas.setMapTool(self.splitTool)

    def manualTileExt(self):
        if not self.actionManualTileExt.isChecked():
            self.setPanMode()
            return
        else:
            self.mapCanvas.setMapTool(self.manualTileTool)

    def manualTileLoft(self):
        if not self.actionManualTileLoft.isChecked():
            self.setPanMode()
            return
        else:
            self.mapCanvas.setMapTool(self.manualLoftTileTool)

    def geometryAdded(self):
        """ Called by map tool when a geometry has been added.
        """
        self.modified = True
        self.mapCanvas.refresh()
        self.actionAdd.setChecked(False)
        self.setPanMode()

    def onGeometryChanged(self):
        """Called by map tool when a geometry has been modified.
        """
        self.modified = True
        self.mapCanvas.refresh()

    def geometryDeleted(self):
        """ Called by map tool when a geometry has been deleted.
        """

        self.modified = True
        self.mapCanvas.refresh()
        self.actionDelete.setChecked(False)
        self.setPanMode()

    def onGetInfo(self, data, GetInfoTool):
        """ Display information about the given track.

            This is called by the "GetInfo" tool when the user clicks on a
            track.
        """

        dialog = HouseInfoDialog(parent=self, data=data, GetInfoTool=GetInfoTool,
                                 conservation_area_list=self.conservation_area_list, curEditedLayer=self.curEditedLayer,
                                 lineLayer=self.lineLayer)
        dialog.loadAttributes()
        layer = getattr(self, 'houseLayer%s' % self.layer_num)

        if dialog.exec():
            dialog.saveAttributes()
            layer.updateFeature(dialog.feature)
            self.modified = True
            self.mapCanvas.refresh()

    def onSplitHouse(self, data, SplitTool):
        """ Split house in quadrilaterals
        """
        gt = geometrytools()
        dialog = SplitHouseDialog(parent=self, data=data, SplitTool=SplitTool, gt=gt)

        if dialog.exec():
            self.modified = True
            self.mapCanvas.refresh()

    def onManualTile(self, data, ManualTile):
        """ Manual tilings
        """

        dialog = ManualTileDialog(parent=self, data=data, ManualTile=ManualTile)

        if dialog.exec():
            self.modified = True
            self.mapCanvas.refresh()

    def onManualLoftTile(self, data, ManualLoftTile):
        """ Manual tilings
        """

        dialog = ManualLoftTileDialog(parent=self, data=data, ManualLoftTile=ManualLoftTile)

        if dialog.exec():
            self.modified = True
            self.mapCanvas.refresh()

    def onHouseAdded(self):
        """ Respond to the user adding a new track.
        """

        self.modified = True
        self.mapCanvas.refresh()
        self.actionAdd.setChecked(False)
        self.setPanMode()

    def onGeometryAdded(self):
        """ Respond to the user adding a new track.
        """

        self.modified = True
        self.mapCanvas.refresh()
        self.actionAdd.setChecked(False)
        self.setPanMode()


def main():
    app = QgsApplication([], True)

    # display_monitor = 2  # the number of the monitor you want to display your widget
    #
    # app.setPrefixPath("/usr", True)
    # app.initQgis()
    #
    # QgsApplication.setPrefixPath(os.environ['QGIS_PREFIX'], True)
    # QgsApplication.initQgis()
    #
    # window = HouseBounds()
    # monitor = QDesktopWidget().screenGeometry(display_monitor)
    # # window.move(monitor.left(), monitor.top())
    # # window.showFullScreen()
    # window.show()
    # window.raise_()
    #
    # app.exec()
    # QgsApplication.exitQgis()


if __name__ == "__main__":
    main()
